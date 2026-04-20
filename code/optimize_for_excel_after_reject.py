import os
import time
import datetime
import threading
import queue
import serial
import numpy as np
from collections import deque
import torch
import pyqtgraph as pg
from pyqtgraph.Qt import QtWidgets, QtCore, QtGui
from Network import RDTNet   # 본인 구조에 맞게 import
from Function import *       # 필요 함수만 있으면 됨 (예: normalize_to_0_1)
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pathlib import Path
import uvicorn

# FastAPI 전역 객체 및 공유 데이터
app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# HTML/JS/CSS 위치 등록
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")

# --- 세션 태그 & 저장 경로 (실시간 저장은 사용 안 해도 무방) ---
RUN_TAG   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
SAVE_ROOT = os.path.join(BASE_DIR, "rdm_frames_0327_209_MJ2")
SAVE_DIR  = os.path.join(SAVE_ROOT, RUN_TAG)   # 실행마다 새 폴더
os.makedirs(SAVE_DIR, exist_ok=True)

# --- 로그 설정 (두 줄만 저장) ---
LOG_DIR  = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, f"rdm_{RUN_TAG}.log")

# --- summary Excel 경로 ---
SUMMARY_XLSX_PATH   = r"d:\Downloads\0412데이터_실시간통계분석.xlsx"
SUMMARY_SHEET_NAME  = "summary_0412"

# fall 이후 ignore window 길이 (프레임 단위)
POST_FALL_IGNORE_FRAMES = 20

# --- 오프라인 RDM 재생용 폴더 설정 ---
RDM_SEQ_DIR      = os.path.join(r"d:\Desktop\종합설계3\0327_rdm\rdm_frames_0327_209_MJ2")
REPEAT_SEQUENCE  = False

# 기본값 (walk / fall / none 등에 사용)
CONS_WINDOW_BASE = 6
CONS_THRESH_BASE = 90.0

# sit / stand 전용: 7프레임, 87%
CONS_WINDOW_SIT_STAND = 6
CONS_THRESH_SIT_STAND = 90.0

# squat 전용: 4프레임, 85%
CONS_WINDOW_SQUAT = 6
CONS_THRESH_SQUAT = 90.0

# fall 전용: 5프레임, 90%
CONS_WINDOW_FALL = 6
CONS_THRESH_FALL = 90.0

# 디버깅 플래그
DEBUG_FSM = True

def append_summary_row_to_excel(
    seq_id,
    duration_sec,
    activity_counts,
    durations_by_cat=None,
    xlsx_path=SUMMARY_XLSX_PATH,
    sheet_name=SUMMARY_SHEET_NAME,
):
    """
    summary 시트에 한 줄(seq 기준) append.
    - seq_id          : RUN_TAG 같은 문자열
    - duration_sec    : 전체 실험 시간(sec)  (2열)
    - activity_counts : {'sit':N, 'stand':N, 'walk_toward':N, 'walk_away':N, 'squat':N, 'fall':N}
    - durations_by_cat: {'sit':sec, 'stand':sec, 'walk':sec}  (열 73~75)
    """
    if not xlsx_path:
        return
    if not os.path.exists(xlsx_path):
        print(f"[WARN] summary Excel을 찾을 수 없어 건너뜀: {xlsx_path}")
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[WARN] openpyxl 미설치 → Excel summary 갱신 생략 (pip install openpyxl)")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        print(f"[WARN] Excel 파일이 열려 있어 쓸 수 없음 (Permission denied): {xlsx_path}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"[WARN] '{sheet_name}' 시트를 찾을 수 없어 건너뜀")
        return
    ws = wb[sheet_name]

    # 1) 첫 데이터 입력 행 찾기 (3행부터: 1행 header, 2행 class 이름행)
    start_row = 3
    target_row = None
    for r in range(start_row, ws.max_row + 5):
        # A~E(1~5열)가 전부 비어 있으면 "빈 행"으로 간주
        if all(ws.cell(row=r, column=c).value is None for c in range(1, 6)):
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1

    # 2) 기본 정보 (seq, duration, no. of activities)
    ws.cell(row=target_row, column=1).value = seq_id
    ws.cell(row=target_row, column=2).value = int(round(float(duration_sec)))

    total_acts = int(
        activity_counts.get("sit", 0)
        + activity_counts.get("stand", 0)
        + activity_counts.get("walk_away", 0)
        + activity_counts.get("walk_toward", 0)
        + activity_counts.get("squat", 0)
        + activity_counts.get("fall", 0)
    )
    ws.cell(row=target_row, column=3).value = total_acts  # "no. of activities"

    # 3) activities(모델) 블록 (열 J~O, 10~15열: sit, stand, walk_a, walk_t, squat, fall)
    col_map = {
        "sit":         10,
        "stand":       11,
        "walk_away":   12,
        "walk_toward": 13,
        "squat":       14,
        "fall":        15,
    }
    for label, col in col_map.items():
        ws.cell(row=target_row, column=col).value = int(activity_counts.get(label, 0))

    # 4) duration(detected) 블록 (열 73~75: sit, stand, walk)
    if durations_by_cat is not None:
        ws.cell(row=target_row, column=73).value = float(durations_by_cat.get("sit", 0.0))
        ws.cell(row=target_row, column=74).value = float(durations_by_cat.get("stand", 0.0))
        ws.cell(row=target_row, column=75).value = float(durations_by_cat.get("walk", 0.0))

    wb.save(xlsx_path)
    print(f"[INFO] Excel summary 갱신: row={target_row}, file={xlsx_path}")

def debug_print(msg: str):
    if DEBUG_FSM:
        print(msg)

# 정적 파일 제공
app.mount("/static", StaticFiles(directory=STATIC_DIR, html=True), name="static")

# ————————————————————————————————
# HTML 라우팅
# ————————————————————————————————
@app.get("/")
def get_splash():
    return FileResponse(os.path.join(STATIC_DIR, "splash.html"))

@app.get("/main.html")
def get_main():
    return FileResponse(os.path.join(STATIC_DIR, "main.html"))

@app.get("/summary.html")
def get_summary():
    return FileResponse(os.path.join(STATIC_DIR, "summary.html"))

# ————————————————————————————————
# status API
# ————————————————————————————————
@app.get("/status")
def get_status():
    return {**status_data, "durations": _durations_snapshot()}

# --- 설정 ---
CONFIG_FILE    = "64x64x10_sub.cfg"   # 지금은 안 써도 됨 (UART 비활성)
MODEL_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'model', 'RDTNet_model52.pth')
DEVICE         = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
RDM_SHAPE      = (64, 64)
WINDOW         = 24  # sliding window 길이
CLASS_NAMES    = ['fall', 'walk_away', 'walk_toward', 'squat', 'sit', 'stand', 'none']
FRAME_INTERVAL = 0.125  # 8 FPS

EVENT_LABELS = ("sit", "stand", "walk_away", "walk_toward", "squat", "fall")
EVENT_LABEL_INDICES = {name: CLASS_NAMES.index(name) for name in EVENT_LABELS}

stand_idx = CLASS_NAMES.index('stand')
squat_idx = CLASS_NAMES.index('squat')

status_data = {
    "squat_count": 0,
    "stand_count": 0,
    "current_action": "none",
    "class_probs": {
        "none": 1.0, "walk_away": 0.0, "walk_toward": 0.0,
        "squat": 0.0, "sit": 0.0, "stand": 0.0, "fall": 0.0
    }
}

# === duration & state (● frame 기반, span 전체를 합산) ===
durations = {
    "walk_sec":        0.0,
    "walk_toward_sec": 0.0,
    "walk_away_sec":   0.0,
    "sit_sec":         0.0,
    "stand_sec":       0.0,
}

span_active      = None        # 'walk' | 'sit' | 'stand' | None
span_start_frame = None        # 해당 span이 시작된 frame index
last_walk_kind   = None        # 'walk_toward' | 'walk_away' | None

LABEL2CAT = {
    "walk_toward": "walk",
    "walk_away":   "walk",
    "sit":         "sit",
    "stand":       "stand",
}

global_frame_idx = 0

def _add_elapsed_by_frames(cat: str, start_frame: int, end_frame: int):
    """
    frame index 기준으로 [start_frame, end_frame) 구간 전체를 cat duration에 누적.
    """
    if cat not in ("walk", "sit", "stand"):
        return
    if start_frame is None or end_frame is None:
        return
    el_frames = int(end_frame) - int(start_frame)
    if el_frames <= 0:
        return

    el_sec = el_frames * FRAME_INTERVAL

    if cat == "walk":
        durations["walk_sec"] += el_sec
        if last_walk_kind == "walk_toward":
            durations["walk_toward_sec"] += el_sec
        elif last_walk_kind == "walk_away":
            durations["walk_away_sec"] += el_sec
    elif cat == "sit":
        durations["sit_sec"] += el_sec
    elif cat == "stand":
        durations["stand_sec"] += el_sec

def _durations_snapshot():
    """
    지금까지 누적된 duration(sec)을 반환.
    - 이미 확정된 span들은 durations[]에 들어가 있고,
    - 현재 진행 중인 span이 있으면, span_start_frame ~ global_frame_idx 까지의 프레임을 추가 반영.
    """
    snap = {
        "walk_sec":        durations.get("walk_sec", 0.0),
        "walk_toward_sec": durations.get("walk_toward_sec", 0.0),
        "walk_away_sec":   durations.get("walk_away_sec", 0.0),
        "sit_sec":         durations.get("sit_sec", 0.0),
        "stand_sec":       durations.get("stand_sec", 0.0),
    }

    if span_active in ("walk", "sit", "stand") and span_start_frame is not None:
        el_frames = int(global_frame_idx) - int(span_start_frame)
        if el_frames > 0:
            el_sec = el_frames * FRAME_INTERVAL
            if span_active == "walk":
                snap["walk_sec"] += el_sec
                if last_walk_kind == "walk_toward":
                    snap["walk_toward_sec"] += el_sec
                elif last_walk_kind == "walk_away":
                    snap["walk_away_sec"] += el_sec
            elif span_active == "sit":
                snap["sit_sec"] += el_sec
            elif span_active == "stand":
                snap["stand_sec"] += el_sec

    return snap

# === 옵션: 비동기 저장기 (지금은 안 써도 됨) ===
SAVE_Q = queue.Queue(maxsize=512)

def save_worker():
    while True:
        item = SAVE_Q.get()
        if item is None:
            SAVE_Q.task_done()
            break
        path, arr = item
        try:
            np.save(path, arr)
        finally:
            SAVE_Q.task_done()

# --- UART 설정/파싱 (오프라인에서는 미사용) ---
def serialConfig(cfg):
    CLIport  = serial.Serial('COM4', 115200)
    Dataport = serial.Serial('COM3', 921600)
    with open(cfg) as f:
        for line in f:
            CLIport.write((line.strip() + '\n').encode())
            time.sleep(0.01)
    return CLIport, Dataport

def parseConfigFile(cfg):
    configParameters = {}
    with open(cfg) as f:
        for line in f:
            sp = line.strip().split()
            if not sp:
                continue
            if sp[0] == "profileCfg":
                startFreq        = float(sp[2])
                idleTime         = float(sp[3])
                rampEndTime      = float(sp[5])
                freqSlopeConst   = float(sp[8])
                numAdcSamples    = int(sp[10])
                digOutSampleRate = int(sp[11])
            elif sp[0] == "frameCfg":
                chirpStartIdx = int(sp[1])
                chirpEndIdx   = int(sp[2])
                numLoops      = int(sp[3])
    numChirps = (chirpEndIdx - chirpStartIdx + 1) * numLoops
    configParameters["numDopplerBins"]       = numChirps
    configParameters["numRangeBins"]         = numAdcSamples
    configParameters["rangeIdxToMeters"]     = (3e8 * digOutSampleRate * 1e3) / (
        2 * freqSlopeConst * 1e12 * numAdcSamples
    )
    configParameters["dopplerResolutionMps"] = 3e8 / (
        2 * startFreq * 1e9 * (idleTime + rampEndTime) * 1e-6 * numChirps
    )
    return configParameters

def readAndParseData18xx(Dataport, configParameters):
    return None

# --- 모델 로드 ---
def load_model(path):
    model = RDTNet([[1,8,16,32,64,128],[128,128,128,128]])
    ckpt  = torch.load(path, map_location=DEVICE)
    model.load_state_dict(ckpt)
    model.to(DEVICE)
    model.eval()
    return model

def main():
    global span_active, span_start_frame, last_walk_kind, global_frame_idx
    global fall_episode_open, post_fall_ignore_remain

    use_async_save = False
    if use_async_save:
        saver_t = threading.Thread(target=save_worker, daemon=True)
        saver_t.start()

    def run_fastapi():
        uvicorn.run(app, host="127.0.0.1", port=8000)

    threading.Thread(target=run_fastapi, daemon=True).start()

    if not os.path.isdir(RDM_SEQ_DIR):
        print(f"[ERROR] RDM_SEQ_DIR가 존재하지 않습니다: {RDM_SEQ_DIR}")
        return

    npy_files = sorted(
        os.path.join(RDM_SEQ_DIR, f)
        for f in os.listdir(RDM_SEQ_DIR)
        if f.lower().endswith(".npy")
    )
    if not npy_files:
        print(f"[ERROR] .npy 파일을 찾을 수 없습니다: {RDM_SEQ_DIR}")
        return

    print(f"[INFO] 재생할 .npy 프레임 개수: {len(npy_files)} (dir={RDM_SEQ_DIR})")

    model = load_model(MODEL_PATH)

    detection_history = deque(maxlen=CONS_WINDOW_SIT_STAND)

    frame_buffer     = deque(maxlen=WINDOW)
    timestamp_buffer = deque(maxlen=WINDOW)
    frame_idx        = 0
    file_idx         = 0

    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] 모델 추론 시작 (윈도우 크기={WINDOW})...")

    log_fp = open(LOG_PATH, "a", encoding="utf-8", buffering=1)

    activity_counts = {name: 0 for name in CLASS_NAMES}

    cons_hold_active = False
    cons_hold_cls    = None

    def log_two_lines(ts_line: str, probs_line: str):
        print(ts_line)
        print(probs_line)
        try:
            log_fp.write(ts_line + "\n")
            log_fp.write(probs_line + "\n")
        except Exception:
            pass

    fall_episode_open       = False
    post_fall_ignore_remain = 0

    sit_recently_counted_wait_sit_sep = False
    squat_confirmed_wait_squat_sep    = False

    try:
        while True:
            # --- 1) 다음 RDM 프레임(.npy) 로드 ---
            if file_idx >= len(npy_files):
                if REPEAT_SEQUENCE:
                    file_idx = 0
                    print("[INFO] .npy 프레임 끝까지 재생 → 처음부터 반복")
                else:
                    print("[INFO] 모든 .npy 프레임 재생 완료. 메인 루프 종료.")
                    break

            fpath = npy_files[file_idx]
            file_idx += 1
            frame_idx += 1
            global_frame_idx = frame_idx

            rdm = np.load(fpath).astype(np.float32)
            if rdm.shape != RDM_SHAPE:
                rdm = rdm[:RDM_SHAPE[0], :RDM_SHAPE[1]]

            frame_ts = time.time()
            frame_buffer.append(rdm)
            timestamp_buffer.append(frame_ts)

            # --- 2) 윈도우가 채워지면 모델 inference ---
            if len(frame_buffer) == WINDOW:
                input_tensor = RDM_prepare(frame_buffer, DEVICE)
                with torch.no_grad():
                    probs = model(input_tensor).cpu().numpy().squeeze(0) * 100

                idx = np.argmax(probs)
                cls = CLASS_NAMES[idx]

                now_ms   = datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]
                ts_line  = f"[{now_ms}] Frame {frame_idx:4d}: {cls:12} ({probs[idx]:.2f}%)"
                probs_str = ", ".join([f"{name}: {p:.1f}%" for name, p in zip(CLASS_NAMES, probs)])
                prob_line = f"  Probabilities: {probs_str}"
                log_two_lines(ts_line, prob_line)

                # ==========================================================
                # [NEW] "나왔다(cls)" 기준 separator 처리 + sit->squat 취소
                # ==========================================================
                if cls == "squat" and sit_recently_counted_wait_sit_sep:
                    if activity_counts.get("sit", 0) > 0:
                        activity_counts["sit"] -= 1
                    debug_print(
                        f"[CANCEL] squat appeared at frame {frame_idx} without non-sit separator "
                        f"after confirmed sit → undo sit count, sit_count={activity_counts.get('sit', 0)}"
                    )
                    sit_recently_counted_wait_sit_sep = False

                elif sit_recently_counted_wait_sit_sep and cls != "sit":
                    sit_recently_counted_wait_sit_sep = False

                debug_print(
                    f"[STATE] frame={frame_idx} cls={cls}({probs[idx]:.1f}%) "
                    f"fall_episode_open={fall_episode_open} "
                    f"ignore_remain={post_fall_ignore_remain} "
                    f"hist_len={len(detection_history)}"
                )

                # post-fall ignore 적용
                append_to_hist = True
                if post_fall_ignore_remain > 0:
                    append_to_hist = False
                    post_fall_ignore_remain -= 1
                    debug_print(
                        f"[IGNORE] frame {frame_idx} after confirmed fall→none "
                        f"(remain={post_fall_ignore_remain})"
                    )

                if append_to_hist:
                    # 클래스가 바뀌면 hold 해제
                    if cons_hold_active and cons_hold_cls is not None and cls != cons_hold_cls:
                        debug_print(
                            f"[CONS] frame {frame_idx} cls changed {cons_hold_cls} -> {cls}, release hold"
                        )
                        cons_hold_active = False
                        cons_hold_cls    = None

                    detection_history.append((cls, probs[idx]))

                    confirmed_cls = None
                    used_window   = None
                    used_classes  = None
                    used_probs    = None

                    # 1) sit / stand
                    if len(detection_history) >= CONS_WINDOW_SIT_STAND and confirmed_cls is None:
                        window    = list(detection_history)[-CONS_WINDOW_SIT_STAND:]
                        w_classes = [c for c, _ in window]
                        w_probs   = [p for _, p in window]
                        if len(set(w_classes)) == 1 and w_classes[0] in ("sit", "stand"):
                            if all(p >= CONS_THRESH_SIT_STAND for p in w_probs):
                                confirmed_cls = w_classes[0]
                                used_window   = CONS_WINDOW_SIT_STAND
                                used_classes  = w_classes
                                used_probs    = w_probs

                    # 2) squat
                    if len(detection_history) >= CONS_WINDOW_SQUAT and confirmed_cls is None:
                        window    = list(detection_history)[-CONS_WINDOW_SQUAT:]
                        w_classes = [c for c, _ in window]
                        w_probs   = [p for _, p in window]
                        if len(set(w_classes)) == 1 and w_classes[0] == "squat":
                            if all(p >= CONS_THRESH_SQUAT for p in w_probs):
                                confirmed_cls = "squat"
                                used_window   = CONS_WINDOW_SQUAT
                                used_classes  = w_classes
                                used_probs    = w_probs

                    # 3) fall
                    if len(detection_history) >= CONS_WINDOW_FALL and confirmed_cls is None:
                        window    = list(detection_history)[-CONS_WINDOW_FALL:]
                        w_classes = [c for c, _ in window]
                        w_probs   = [p for _, p in window]
                        if len(set(w_classes)) == 1 and w_classes[0] == "fall":
                            if all(p >= CONS_THRESH_FALL for p in w_probs):
                                confirmed_cls = "fall"
                                used_window   = CONS_WINDOW_FALL
                                used_classes  = w_classes
                                used_probs    = w_probs

                    # 4) 나머지(walk_*, none)
                    if len(detection_history) >= CONS_WINDOW_BASE and confirmed_cls is None:
                        window    = list(detection_history)[-CONS_WINDOW_BASE:]
                        w_classes = [c for c, _ in window]
                        w_probs   = [p for _, p in window]
                        if len(set(w_classes)) == 1:
                            candidate = w_classes[0]
                            if candidate not in ("sit", "stand", "squat", "fall"):
                                if all(p >= CONS_THRESH_BASE for p in w_probs):
                                    confirmed_cls = candidate
                                    used_window   = CONS_WINDOW_BASE
                                    used_classes  = w_classes
                                    used_probs    = w_probs

                    # ==========================================================
                    # ✅ confirmed_cls 처리 (COUNT 규칙 + DURATION FSM 복구 + UI 복구)
                    # ==========================================================
                    if confirmed_cls is not None:
                        # 1) 새 이벤트인가? (confirmed 기준)
                        is_new_event = (not cons_hold_active) or (cons_hold_cls != confirmed_cls)

                        # ----------------------------
                        # (A) fall 오검출 취소 (예전 로직 복구)
                        # ----------------------------
                        if fall_episode_open and confirmed_cls not in ("fall", "none"):
                            if activity_counts.get("fall", 0) > 0:
                                activity_counts["fall"] -= 1
                            fall_episode_open = False
                            post_fall_ignore_remain = 0

                        # ----------------------------
                        # [B] ✅ DURATION FSM (confirmed 기준)  <<<< 여기 복구된 핵심
                        #     - 카운트 억제(do_count=False)여도 duration은 "확정된 상태"를 따르도록 함
                        # ----------------------------
                        current_frame = frame_idx
                        new_cat = LABEL2CAT.get(confirmed_cls, None)

                        if new_cat in ("walk", "sit", "stand"):
                            if span_active is None:
                                span_active      = new_cat
                                span_start_frame = current_frame
                                if new_cat == "walk":
                                    last_walk_kind = confirmed_cls
                            else:
                                if new_cat == span_active:
                                    if new_cat == "walk":
                                        last_walk_kind = confirmed_cls
                                else:
                                    _add_elapsed_by_frames(span_active, span_start_frame, current_frame)
                                    span_active      = new_cat
                                    span_start_frame = current_frame
                                    if new_cat == "walk":
                                        last_walk_kind = confirmed_cls

                        elif confirmed_cls == "none":
                            # none은 span 유지(닫지 않음)
                            pass

                        else:
                            # fall, squat 등 → 진행 중이던 walk/sit/stand span 종료
                            if span_active in ("walk", "sit", "stand"):
                                _add_elapsed_by_frames(span_active, span_start_frame, current_frame)
                                span_active      = None
                                span_start_frame = None

                        # ----------------------------
                        # [C] COUNT/규칙은 "새 이벤트"일 때만 적용 (네 기존 의도 유지)
                        # ----------------------------
                        if is_new_event:
                            do_count = True

                            # (B) squat 직후 stand 확정은 separator 전까지 카운트 금지
                            if confirmed_cls == "stand" and squat_confirmed_wait_squat_sep:
                                do_count = False

                            # (C) sit->(separator 없이) squat이면 직전 sit 취소 (확정 기준)
                            if confirmed_cls == "squat" and sit_recently_counted_wait_sit_sep:
                                if activity_counts.get("sit", 0) > 0:
                                    activity_counts["sit"] -= 1
                                sit_recently_counted_wait_sit_sep = False

                            # 2) 카운트
                            if do_count and confirmed_cls in ("sit","stand","walk_toward","walk_away","squat","fall"):
                                activity_counts[confirmed_cls] += 1

                            # 3) hold 갱신
                            cons_hold_active = True
                            cons_hold_cls = confirmed_cls

                            # 4) sit separator 플래그
                            if confirmed_cls == "sit":
                                sit_recently_counted_wait_sit_sep = True
                            elif confirmed_cls != "sit":
                                sit_recently_counted_wait_sit_sep = False

                            # 5) squat separator 플래그
                            if confirmed_cls == "squat":
                                squat_confirmed_wait_squat_sep = True
                            elif confirmed_cls not in ("squat", "stand"):
                                squat_confirmed_wait_squat_sep = False

                            # 6) fall 에피소드 / ignore window
                            if confirmed_cls == "fall":
                                fall_episode_open = True
                            elif confirmed_cls == "none" and fall_episode_open:
                                fall_episode_open = False
                                post_fall_ignore_remain = POST_FALL_IGNORE_FRAMES
                                detection_history.clear()
                                cons_hold_active = False
                                cons_hold_cls = None

                        # ----------------------------
                        # [D] UI(status_data) 업데이트 복구
                        # ----------------------------
                        status_data["current_action"] = confirmed_cls
                        status_data["class_probs"] = {
                            name: round(float(p/100), 3) for name, p in zip(CLASS_NAMES, probs)
                        }

                        status_data["squat_count"] = int(activity_counts.get("squat", 0))
                        status_data["stand_count"] = int(activity_counts.get("stand", 0))
                        status_data["sit_count"]         = int(activity_counts.get("sit", 0))
                        status_data["walk_away_count"]   = int(activity_counts.get("walk_away", 0))
                        status_data["walk_toward_count"] = int(activity_counts.get("walk_toward", 0))
                        status_data["fall_count"]        = int(activity_counts.get("fall", 0))

                        debug_print(
                            f"[CONFIRM] frame {frame_idx} → {confirmed_cls} "
                            f"(win={used_window}) "
                            f"via window classes={used_classes} "
                            f"probs={[f'{p:.1f}' for p in (used_probs or [])]} "
                            f"counts={activity_counts} "
                            f"dur={_durations_snapshot()}"
                        )

    except KeyboardInterrupt:
        print("\n[INFO] 키보드 인터럽트: 프로그램 종료")

    # === 로그 종료 후 Excel summary 업데이트 ===
    try:
        total_run_sec = frame_idx * FRAME_INTERVAL

        dur_snap = _durations_snapshot()
        durations_by_cat = {
            "sit":   float(dur_snap.get("sit_sec", 0.0)),
            "stand": float(dur_snap.get("stand_sec", 0.0)),
            "walk":  float(dur_snap.get("walk_sec", 0.0)),
        }

        append_summary_row_to_excel(
            seq_id=RUN_TAG,
            duration_sec=total_run_sec,
            activity_counts=activity_counts,
            durations_by_cat=durations_by_cat,
        )
    except Exception as e:
        print(f"[WARN] Excel summary 갱신 실패: {e}")

    try:
        log_fp.close()
    except:
        pass

    if use_async_save:
        try:
            SAVE_Q.put(None)
            SAVE_Q.join()
        except:
            pass

if __name__ == "__main__":
    main()
